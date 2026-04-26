import os
import pytz
from datetime import datetime, timedelta
import openpyxl
import telebot
from telebot import types
import threading
import time

# Configuration
TELEGRAM_TOKEN = 'Не покажу :)'
EXCEL_FILE = 'Таблица пример уроки.xlsx'  # Локальный файл Excel
timeDelay = [20, 19]

# Initialize bot
bot = telebot.TeleBot(TELEGRAM_TOKEN)

# Timezone setup
MOSCOW_TZ = pytz.timezone('Europe/Moscow')

# User states and data
user_states = {}
user_data = {}


# Load Excel file
def load_excel_file():
    return openpyxl.load_workbook(EXCEL_FILE)


# Get schedule data
def get_schedule_data(wb):
    schedule_sheet = wb['расписание']
    data = []
    for row in schedule_sheet.iter_rows(values_only=True):
        data.append(row)
    return data


# Get students data
def get_students_data(wb):
    students_sheet = wb['ученики']
    data = []
    for row in students_sheet.iter_rows(values_only=True):
        data.append(row)
    return data


# Get teachers data
def get_teachers_data(wb):
    teachers_sheet = wb['Преподаватели']
    data = []
    for row in teachers_sheet.iter_rows(values_only=True):
        data.append(row)
    return data


# Parse schedule (аналогично предыдущей версии)
def parse_schedule(schedule_data):
    lessons = []
    current_day = None
    current_time = None

    for row in schedule_data:
        # Новый способ определения дня недели
        for cell in row:
            if str(cell).strip() in ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']:
                current_day = str(cell).strip()
                break

        if ':' in str(row[0]):  # Строка с временем
            current_time = row[0]
            day_names = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']

            for i in range(1, 8):  # Колонки B-H
                if i < len(row) and row[i] and '[' in str(row[i]):
                    cell_content = str(row[i])
                    student_info = cell_content.split('[')
                    name = student_info[0].strip()
                    student_id = student_info[1].replace(']', '').strip()

                    lesson_time = str(current_time).split('-')[0].strip()
                    try:
                        hour, minute = map(int, lesson_time.split(':'))
                        lesson_datetime = datetime.now(MOSCOW_TZ).replace(
                            hour=hour, minute=minute, second=0, microsecond=0
                        )

                        # Получаем день недели по индексу столбца
                        current_day = day_names[i - 1]  # i = 1 → 'ПН', i = 2 → 'ВТ', и т.д.
                        target_weekday = day_names.index(current_day)

                        while lesson_datetime.weekday() != target_weekday:
                            lesson_datetime += timedelta(days=1)

                        lessons.append({
                            'day': current_day,
                            'time': current_time,
                            'datetime': lesson_datetime,
                            'name': name,
                            'id': student_id
                        })
                    except (ValueError, IndexError):
                        continue

    return lessons


# Find user by ID
def find_user_by_id(user_id, users_data):
    for row in users_data[1:]:  # Пропускаем заголовок
        if len(row) > 2 and str(row[2]) == str(user_id):
            return {
                'name': row[1],
                'timezone': row[3] if len(row) > 3 else 'мск',
                'parent_name': row[4] if len(row) > 4 else '',
                'parent_contact': row[5] if len(row) > 5 else '',
                'student_contact': row[6] if len(row) > 6 else ''
            }
    return None


# Обработчики команд бота (аналогично предыдущей версии)
@bot.message_handler(commands=['start'])
def handle_start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Я ученик'), types.KeyboardButton('Я преподаватель'))
    bot.send_message(message.chat.id, "Выберите вашу роль:", reply_markup=markup)
    user_states[message.chat.id] = 'awaiting_role'

# старый код, работает.
@bot.message_handler(commands=['my_lessons'])
def handle_my_lessons(message):
    chat_id = message.chat.id

    if chat_id not in user_data or 'id' not in user_data[chat_id]:
        bot.send_message(chat_id, "Вы ещё не зарегистрировались. Введите /start.")
        return

    user_id = user_data[chat_id]['id']
    role = user_data[chat_id]['role']
    wb = load_excel_file()
    schedule_data = get_schedule_data(wb)
    lessons = parse_schedule(schedule_data)

    relevant_lessons = [l for l in lessons if str(l['id']) == str(user_id)]
    if not relevant_lessons:
        bot.send_message(chat_id, "Нет занятий, связанных с вами.")
        return

    message_text = "📅 Ваше расписание:\n"
    for lesson in relevant_lessons:
        if role == 'student':
            message_text += f"• {lesson['day']} в {lesson['time']}\n"
        else:  # teacher
            message_text += f"• {lesson['day']} в {lesson['time']} — с {lesson['name']}\n"

    bot.send_message(chat_id, message_text)



@bot.message_handler(func=lambda message: True)
def handle_message(message):
    chat_id = message.chat.id

    if message.text == "Показать расписание":
        if chat_id not in user_data or 'id' not in user_data[chat_id]:
            bot.send_message(chat_id, "Вы ещё не зарегистрировались. Введите /start.")
            return

        user_id = user_data[chat_id]['id']
        role = user_data[chat_id]['role']
        wb = load_excel_file()
        schedule_data = get_schedule_data(wb)
        lessons = parse_schedule(schedule_data)

        relevant_lessons = [l for l in lessons if str(l['id']) == str(user_id)]
        if not relevant_lessons:
            bot.send_message(chat_id, "Нет занятий, связанных с вами.")
            return

        message_text = "📅 Ваше расписание:\n"
        for lesson in relevant_lessons:
            if role == 'student':
                message_text += f"• {lesson['day']} в {lesson['time']}\n"
            else:
                message_text += f"• {lesson['day']} в {lesson['time']} — с {lesson['name']}\n"

        bot.send_message(chat_id, message_text)
        return

    if message.text == "Отключить уведомления":
        user_states.pop(chat_id, None)
        user_data.pop(chat_id, None)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('Я ученик'), types.KeyboardButton('Я преподаватель'))
        bot.send_message(chat_id, "Вы отключили уведомления. Выберите вашу роль:", reply_markup=markup)
        user_states[chat_id] = 'awaiting_role'
        return

    if chat_id in user_states and user_states[chat_id] == 'awaiting_role':
        if message.text in ['Я ученик', 'Я преподаватель']:
            user_data[chat_id] = {'role': 'student' if message.text == 'Я ученик' else 'teacher'}
            bot.send_message(chat_id, "Введите ваш ID (число в квадратных скобках в расписании):")
            user_states[chat_id] = 'awaiting_id'
        else:
            bot.send_message(chat_id, "Пожалуйста, выберите вариант из предложенных.")

    elif chat_id in user_states and user_states[chat_id] == 'awaiting_id':
        try:
            user_id = message.text.strip()
            wb = load_excel_file()

            if user_data[chat_id]['role'] == 'student':
                students_data = get_students_data(wb)
                user_info = find_user_by_id(user_id, students_data)
                if user_info:
                    user_data[chat_id]['id'] = user_id
                    user_data[chat_id]['info'] = user_info
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    markup.add(types.KeyboardButton('Показать расписание'))
                    markup.add(types.KeyboardButton('Отключить уведомления'))
                    bot.send_message(chat_id,
                                     f"Отлично, {user_info['name']}! Вы будете получать уведомления о своих занятиях.",
                                     reply_markup=markup)
                    user_states[chat_id] = 'registered'

                else:
                    bot.send_message(chat_id, "Ученик с таким ID не найден. Попробуйте еще раз.")

            else:  # teacher
                # Для преподавателей просто сохраняем ID
                user_data[chat_id]['id'] = user_id
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add(types.KeyboardButton('Показать расписание'))
                markup.add(types.KeyboardButton('Отключить уведомления'))
                bot.send_message(chat_id,
                                 "Вы зарегистрированы как преподаватель. Вы будете получать уведомления о ваших занятиях.",
                                 reply_markup=markup)
                user_states[chat_id] = 'registered'


        except Exception as e:
            bot.send_message(chat_id, f"Произошла ошибка: {str(e)}. Попробуйте еще раз.")


# Проверка предстоящих занятий
def check_upcoming_lessons():
    while True:
        try:
            now = datetime.now(MOSCOW_TZ)
            wb = load_excel_file()
            schedule_data = get_schedule_data(wb)
            lessons = parse_schedule(schedule_data)

            for lesson in lessons:
                for delay in timeDelay:
                    reminder_time = lesson['datetime'] - timedelta(minutes=delay)

                    if now <= reminder_time <= now + timedelta(minutes=1):
                        for chat_id, data in user_data.items():
                            if 'id' in data and str(data['id']) == str(lesson['id']):
                                if data['role'] == 'student':
                                    message = f"🔔 Напоминание: у вас занятие {lesson['day']} в {lesson['time']} (через {delay} мин)"
                                else:
                                    message = f"🔔 Напоминание: у вас занятие с {lesson['name']} {lesson['day']} в {lesson['time']} (через {delay} мин)"

                                try:
                                    bot.send_message(chat_id, message)
                                except Exception as e:
                                    print(f"Не удалось отправить сообщение {chat_id}: {e}")

            time.sleep(30)  # Проверяем каждые 30 секунд

        except Exception as e:
            print(f"Ошибка в check_upcoming_lessons: {e}")
            time.sleep(60)


# Запускаем проверку занятий в отдельном потоке
reminder_thread = threading.Thread(target=check_upcoming_lessons)
reminder_thread.daemon = True
reminder_thread.start()

# Запускаем бота
print("Бот запущен...")
bot.polling()
